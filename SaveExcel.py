
import os
import time
from openpyxl import load_workbook
def update_sheet_preserving_format(excel_file, sheet_name, df):
    # Load the workbook and select the sheet
    book = load_workbook(excel_file)
    sheet = book[sheet_name]

    # Assuming the first row is the header and data starts from the second row
    row_offset = 2  # Start updating from this Excel row

    # Iterate over the DataFrame and update cell values
    for r, (index, row) in enumerate(df.iterrows(), start=row_offset):
        for c, col in enumerate(df.columns, start=1):  # Assuming data starts from the first column
            cell = sheet.cell(row=r, column=c)
            cell.value = row[col]

    # Attempt to save the workbook
    try:
        book.save(excel_file)
    except PermissionError:
        # If the file is open and cannot be written to, save with a new name
        base, extension = os.path.splitext(excel_file)
        new_filename = f"{base}_{time.strftime('%Y-%m-%d_%H-%M-%S')}{extension}"
        book.save(new_filename)
        print(f"Original file was open; changes were saved to '{new_filename}' instead.")